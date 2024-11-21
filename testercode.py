import streamlit as st
import os
import shutil
import base64
import fitz
import tempfile
import datetime
import pandas as pd
import csv

st.set_page_config(
    page_title="QTG Review System",
    page_icon="📖",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.sidebar.header("QTG Review System")
st.sidebar.markdown("Manage and review your QTG files efficiently.")
device = st.sidebar.selectbox("Select Device", options=["Select", "FFS", "FTD"], key="device_selection")
year = st.sidebar.selectbox("Select Year", options=["Select", "2023", "2024"], key="year_selection")
set = st.sidebar.selectbox("Select Set", options=["Select", "Set A", "Set B", "Set C", "Set D"], key="set_selection")
if device == "Select" or year == "Select" or set == "Select":
    st.sidebar.warning("Please select a device, year, and set to proceed.")
    st.write("### 🚫 Please select a device, year, and respective QTG set in the sidebar to continue.")
    st.stop()  

base_folder = os.path.join(f"./{device}", year, set)
source_folder = os.path.join(base_folder, "source_folder")
pass_folder = os.path.join(base_folder, "pass_folder")
fail_folder = os.path.join(base_folder, "fail_folder")
signed_folder = os.path.join(base_folder, "signed_folder")
log_file = os.path.join(base_folder, "signing_log.csv")
for folder in [source_folder, pass_folder, fail_folder, signed_folder]:
    if not os.path.exists(folder):
        os.makedirs(folder)

if not os.path.exists(log_file):
    with open(log_file, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Document Name", "Signed By", "Date/Time"])

if "update" not in st.session_state:
    st.session_state["update"] = False

if st.session_state["update"]:
    st.session_state["update"] = False
    st.experimental_rerun()
    
st.write(f"### 📂 QTG Review System: {device} - {year} - {set}")

def list_files(folder):
    """List files in a folder."""
    try:
        return os.listdir(folder)
    except FileNotFoundError:
        return []

def move_file(src, dest):
    try:
        shutil.move(src, dest)
        return True
    except Exception as e:
        st.error(f"Error moving file {src} to {dest}: {e}")
        return False

def retrieve_files(selected_files, src_folder, dest_folder):
    moved_files = []
    for item in selected_files:
        src_path = os.path.join(src_folder, item)
        dest_path = os.path.join(dest_folder, item)
        if move_file(src_path, dest_path):
            moved_files.append(item)
    return moved_files

def get_file_link(file_path):
    with open(file_path, "rb") as f:
        pdf_data = f.read()
        b64 = base64.b64encode(pdf_data).decode()
        return f'<a href="data:application/pdf;base64,{b64}" target="_blank">Open PDF</a>'
    

def update_signing_log(document_name, signer_name, signing_date):
    excel_file = os.path.join(base_folder, "signing_log.xlsx")

    try:
        if not os.path.exists(excel_file):
            # Create the Excel file if it doesn't exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Signing Log"
            ws.append(["Document Name", "Signed By", "Date/Time", "", "", "", "", "", "Signer Name", "Sign Date"])  # Add headers
            wb.save(excel_file)

        # Load the existing Excel file
        wb = load_workbook(excel_file)
        ws = wb.active

        # Find the next available row for I and J columns
        next_row = ws.max_row + 1

        ws[f"I{next_row}"] = signer_name
        ws[f"J{next_row}"] = signing_date

        wb.save(excel_file)
        wb.close()

    except Exception as e:
        st.error(f"Error updating signing log: {e}")

from openpyxl import Workbook, load_workbook

def add_signature_to_pdf(pdf_path, signature_path, signed_folder, signer_name):
    try:
        doc = fitz.open(pdf_path)
        signature_img = fitz.open(signature_path)

        current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Add signature and save the signed PDF
        page = doc[0]  
        top_margin = 20 
        x = 50 
        image_width = 150
        image_height = 150

        image_rect = fitz.Rect(x, top_margin, x + image_width, top_margin + image_height)
        page.insert_image(image_rect, filename=signature_path)

        text_rect = fitz.Rect(x, top_margin + image_height + 5, x + 200, top_margin + image_height + 25)
        page.insert_text(text_rect.tl, f"Signed on: {current_datetime}", fontsize=10)

        signed_file_path = os.path.join(signed_folder, os.path.basename(pdf_path))
        doc.save(signed_file_path)
        doc.close()

        # Update the Excel log
        update_signing_log(os.path.basename(pdf_path), signer_name, current_datetime)

        if os.path.exists(signed_file_path):
            os.remove(pdf_path)

        return signed_file_path

    except Exception as e:
        st.error(f"Error signing PDF: {e}")
        return None


def create_file_dataframe(folder_path):
    folder_contents = list_files(folder_path)
    if folder_contents: 
        data = [[item] for item in folder_contents]
        return pd.DataFrame(data, columns=["File Name"])
    else: 
        return pd.DataFrame(columns=["File Name"])

tab1, tab2, tab3, tab4 = st.tabs(
    ["📂 Manage QTG Files", "🔄 Retrieve Files", "🖊️ E-Sign Document", "📊 Signing Log"]
)

with tab1:
    st.subheader("📁 QTG Review and Classification")

    with st.container():
        files = list_files(source_folder)

        if files:
            file_to_move = st.selectbox("Select the QTG to review", files)
            if file_to_move:
                pdf_file_path = os.path.join(source_folder, file_to_move)
                if os.path.exists(pdf_file_path):
                    with open(pdf_file_path, "rb") as pdf_file:
                        pdf_data = pdf_file.read()
                        st.download_button(
                            label="📂 Open PDF",
                            data=pdf_data,
                            file_name=file_to_move,
                            mime="application/pdf",
                        )
            status = st.radio("Status of QTG", ["Pass", "Fail"], horizontal=True)
            if st.button("Submit"):
                destination_folder = pass_folder if status == "Pass" else fail_folder
                src_path = os.path.join(source_folder, file_to_move)
                dest_path = os.path.join(destination_folder, file_to_move)

                if move_file(src_path, dest_path):
                    st.success(f"Moved '{file_to_move}' to **{'Pass' if status == 'Pass' else 'Fail'}** folder.")
                    st.experimental_rerun()
        else:
            st.warning("🚫 No files to move in the source folder.")

    st.markdown("---")

    with st.container():
        st.subheader("📂 View Folder Contents")

    # Segmented control to select which folder to view
    selected_folder = st.segmented_control(
        label="Choose a folder to view files:",
        options=["Source Folder", "Pass Folder", "Fail Folder"],
        key="folder_selection1",
    )

    # Default message for no selection
    if not selected_folder:
        st.warning("Please select a folder to view its contents.")
    else:
        # Map selected folder to the corresponding path
        folder_map = {
            "Source Folder": source_folder,
            "Pass Folder": pass_folder,
            "Fail Folder": fail_folder,
        }
        current_folder = folder_map[selected_folder]

        # Display contents of the selected folder
        st.markdown(f"### Files in {selected_folder}")
        folder_data = create_file_dataframe(current_folder)
        if not folder_data.empty:
            st.dataframe(folder_data)
        else:
            st.write(f"No files in {selected_folder.lower()}.")

with tab2:
    st.header("🔄 Retrieve Files to Source Folder")

    # Segmented control for folder selection
    selected_folder = st.segmented_control(
        label="Choose folder to view files:",
        options=["Pass Folder", "Fail Folder"],
        key="folder_selection2",
    )

    # Ensure the user selects a folder before proceeding
    if selected_folder:
        folder_path = pass_folder if selected_folder == "Pass Folder" else fail_folder
        folder_files = list_files(folder_path)

        if folder_files:
            # Show files in the selected folder using a radio button
            st.markdown(f"### 📂 Files in {selected_folder}")
            file_to_retrieve = st.radio(
                "Select a file to retrieve:",
                options=folder_files,
                key=f"{selected_folder}_radio",  # Unique key for each folder
            )

            # Retrieve the selected file
            if st.button(f"Retrieve Selected File from {selected_folder}"):
                with st.spinner("Retrieving file..."):
                    moved_files = retrieve_files([file_to_retrieve], folder_path, source_folder)
                if moved_files:
                    st.success(f"Successfully moved '{file_to_retrieve}' to the Source folder!")
                    st.rerun()  # Refresh the interface to update the file list
                else:
                    st.warning("No file was moved.")
        else:
            st.warning(f"No files found in {selected_folder}.")
    else:
        st.info("Please select a folder to view its files.")


with tab3:
    st.header("🖊️ E-Sign Document")
    signature_image = st.file_uploader(
        "Choose an image file for your signature", type=["png", "jpg", "jpeg"], key="signature_uploader"
    )
    signer_name = st.text_input("Enter your name", key="signer_name")
    pass_files = list_files(pass_folder)

    if pass_files:
        file_to_sign = st.selectbox("Select a document to sign", pass_files)
        selected_file_path = os.path.join(pass_folder, file_to_sign)

        if st.button("Apply Signature") and signature_image and signer_name:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(signature_image.read())
                signature_path = tmp.name

            signed_file_path = add_signature_to_pdf(
                selected_file_path, signature_path, signed_folder, signer_name
            )
            if signed_file_path:
                st.success(f"Signed {file_to_sign} successfully!")
                with open(signed_file_path, "rb") as file:
                    st.download_button(
                        "Download Signed Document",
                        data=file,
                        file_name=os.path.basename(signed_file_path),
                        mime="application/pdf",
                    )
    else:
        st.warning("No files available to sign in the Pass folder.")
with tab4:
    st.header("📊 Signing Log")

    excel_file = os.path.join(base_folder, "signing_log.xlsx")

    if os.path.exists(excel_file):
        # Display the signing log from the Excel file
        try:
            log_data = pd.read_excel(excel_file, sheet_name="Signing Log")
            st.dataframe(log_data)
        except Exception as e:
            st.error(f"Error loading signing log: {e}")
    else:
        st.warning("The signing log file does not exist. It will be created when a document is signed.")